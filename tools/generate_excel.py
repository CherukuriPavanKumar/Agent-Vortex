from typing import Annotated
from pathlib import Path
import json

import openpyxl
from openpyxl.styles import Font
from langchain_core.tools import tool


EXCEL_OUTPUT_DIR = Path("generated_excels")


@tool
def generate_excel(
    json_data: Annotated[
        str,
        "JSON string containing a list of dictionaries. Each dictionary represents a row.",
    ],
    output_path: Annotated[
        str,
        "Desired Excel filename. Files are always stored in generated_excels/.",
    ],
    sheet_name: Annotated[
        str,
        "Worksheet name.",
    ] = "Sheet1",
) -> str:
    """
    Generate an Excel (.xlsx) spreadsheet from JSON data.

    Features:
    - Creates generated_excels/ automatically
    - Prevents path traversal
    - Auto-generates headers
    - Handles rows with different fields
    - Auto-sizes columns
    - Verifies workbook creation
    """

    try:
        if not json_data.strip():
            return "Excel generation failed: JSON data is empty."

        parsed_data = json.loads(json_data)

        if not isinstance(parsed_data, list):
            return (
                "Excel generation failed: "
                "Expected a JSON array (list of dictionaries)."
            )

        if len(parsed_data) == 0:
            return "Excel generation failed: Data list is empty."

        if not all(isinstance(row, dict) for row in parsed_data):
            return (
                "Excel generation failed: "
                "Every item in the list must be a dictionary."
            )

        EXCEL_OUTPUT_DIR.mkdir(
            parents=True,
            exist_ok=True,
        )

        filename = Path(output_path).name

        if not filename.lower().endswith(".xlsx"):
            filename += ".xlsx"

        final_path = EXCEL_OUTPUT_DIR / filename

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active

        # Excel sheet name max length = 31
        ws.title = sheet_name[:31]

        # Collect ALL headers from ALL rows
        headers = set()

        for row in parsed_data:
            headers.update(row.keys())

        headers = sorted(headers)

        # Header row
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Data rows
        for row in parsed_data:
            ws.append(
                [
                    row.get(header, "")
                    for header in headers
                ]
            )

        # Auto-size columns
        for column in ws.columns:
            max_length = 0

            column_letter = column[0].column_letter

            for cell in column:
                try:
                    value = str(cell.value)

                    if len(value) > max_length:
                        max_length = len(value)

                except Exception:
                    pass

            ws.column_dimensions[
                column_letter
            ].width = min(max_length + 2, 60)

        wb.save(final_path)

        # Verify file exists
        if not final_path.exists():
            return (
                f"Excel generation failed: "
                f"file was not created at '{final_path}'."
            )

        file_size = final_path.stat().st_size

        if file_size == 0:
            return (
                f"Excel generation failed: "
                f"'{final_path}' is empty (0 bytes)."
            )

        # Verify workbook can actually be reopened
        try:
            openpyxl.load_workbook(final_path)
        except Exception as verify_error:
            return (
                "Excel generation failed: workbook was written "
                f"but could not be reopened. Error: {verify_error}"
            )

        return (
            f"Successfully generated Excel file.\n"
            f"Location: {final_path.resolve()}\n"
            f"Rows: {len(parsed_data)}\n"
            f"Columns: {len(headers)}\n"
            f"Size: {file_size} bytes"
        )

    except json.JSONDecodeError as e:
        return (
            "Excel generation failed: Invalid JSON format.\n"
            f"{e}"
        )

    except Exception as e:
        return f"Excel generation failed: {e}"